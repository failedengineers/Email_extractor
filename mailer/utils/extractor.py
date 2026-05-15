import re
import csv
from openpyxl import load_workbook

EMAIL_REGEX = r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b"


def extract_emails(file_path):
    emails = []

    if file_path.endswith(".csv"):
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                emails.extend(re.findall(EMAIL_REGEX, line))

    elif file_path.endswith(".xlsx"):
        wb = load_workbook(file_path, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        emails.extend(re.findall(EMAIL_REGEX, str(cell)))

    return emails
